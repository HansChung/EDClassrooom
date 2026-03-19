from __future__ import annotations

import csv
import json
from datetime import datetime
from io import BytesIO, StringIO
from zipfile import BadZipFile

from xml.parsers.expat import ExpatError

from flask import Response, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from odf.opendocument import OpenDocumentSpreadsheet, load
from odf import teletype
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models import BookingPeriod, BookingRequest, Classroom, ClassroomBlock, CourseSchedule, Role, SystemRule, User
from app.modules.admin import bp
from app.modules.booking.services import create_booking
from app.modules.notifications.services import create_notification
from app.security import roles_required


REQUIRED_SCHEDULE_COLUMNS = {
    "classroom_code",
    "course_name",
    "instructor_name",
    "weekday",
    "start_time",
    "end_time",
    "semester_label",
    "is_active",
}


def _get_dashboard_context() -> dict:
    rooms = db.session.query(Classroom).order_by(Classroom.code.asc()).all()
    rules = db.session.query(SystemRule).order_by(SystemRule.key.asc()).all()
    users = db.session.query(User).order_by(User.display_name.asc()).limit(100).all()
    schedules = (
        db.session.query(CourseSchedule)
        .join(Classroom)
        .order_by(Classroom.code.asc(), CourseSchedule.weekday.asc(), CourseSchedule.start_time.asc())
        .all()
    )
    blocks = (
        db.session.query(ClassroomBlock)
        .join(Classroom)
        .order_by(ClassroomBlock.start_at.asc(), Classroom.code.asc())
        .all()
    )
    periods = (
        db.session.query(BookingPeriod)
        .order_by(BookingPeriod.sort_order.asc(), BookingPeriod.code.asc())
        .all()
    )
    return {
        "rooms": rooms,
        "rules": rules,
        "users": users,
        "schedules": schedules,
        "blocks": blocks,
        "booking_periods": periods,
        "booking_periods_text": "\n".join(
            f"{period.code},{period.start_time:%H:%M},{period.end_time:%H:%M}"
            for period in periods
        ),
    }


def _normalize_schedule_columns(fieldnames: list[str] | None) -> list[str]:
    return [(name or "").strip() for name in (fieldnames or [])]


def _read_schedule_rows_from_csv(file_storage) -> tuple[list[dict[str, str]], list[str]]:
    content = file_storage.read().decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))
    fieldnames = _normalize_schedule_columns(reader.fieldnames)
    rows = [{(key or "").strip(): (value or "").strip() for key, value in row.items()} for row in reader]
    return rows, fieldnames


def _read_schedule_rows_from_xlsx(file_storage) -> tuple[list[dict[str, str]], list[str]]:
    try:
        workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    except BadZipFile as exc:
        raise ValueError("XLSX 檔案格式有誤。") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = _normalize_schedule_columns([str(cell or "").strip() for cell in rows[0]])
    normalized_rows: list[dict[str, str]] = []
    for values in rows[1:]:
        normalized_rows.append(
            {
                headers[index]: "" if index >= len(values) or values[index] is None else str(values[index]).strip()
                for index in range(len(headers))
            }
        )
    return normalized_rows, headers


def _cell_text(cell: TableCell) -> str:
    return teletype.extractText(cell).strip()


def _read_schedule_rows_from_ods(file_storage) -> tuple[list[dict[str, str]], list[str]]:
    try:
        document = load(BytesIO(file_storage.read()))
    except (BadZipFile, ExpatError, OSError) as exc:
        raise ValueError("ODS 檔案格式有誤。") from exc
    tables = document.spreadsheet.getElementsByType(Table)
    if not tables:
        return [], []
    table = tables[0]
    raw_rows: list[list[str]] = []
    for row in table.getElementsByType(TableRow):
        raw_cells = [_cell_text(cell) for cell in row.getElementsByType(TableCell)]
        if any(raw_cells):
            raw_rows.append(raw_cells)
    if not raw_rows:
        return [], []
    headers = _normalize_schedule_columns(raw_rows[0])
    normalized_rows: list[dict[str, str]] = []
    for values in raw_rows[1:]:
        normalized_rows.append(
            {
                headers[index]: values[index].strip() if index < len(values) else ""
                for index in range(len(headers))
            }
        )
    return normalized_rows, headers


def _read_schedule_rows(file_storage) -> tuple[list[dict[str, str]], list[str]]:
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        return _read_schedule_rows_from_csv(file_storage)
    if filename.endswith(".xlsx"):
        return _read_schedule_rows_from_xlsx(file_storage)
    if filename.endswith(".ods"):
        return _read_schedule_rows_from_ods(file_storage)
    raise ValueError("目前僅支援 CSV、XLSX、ODS 格式。")


def _validate_schedule_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    valid_rows: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=2):
        messages: list[str] = []
        classroom_code = (row.get("classroom_code") or "").strip().upper()
        course_name = (row.get("course_name") or "").strip()
        instructor_name = (row.get("instructor_name") or "").strip()
        semester_label = (row.get("semester_label") or "").strip() or "current"
        is_active_raw = (row.get("is_active") or "").strip().lower()

        classroom = None
        if not classroom_code:
            messages.append("缺少 classroom_code")
        else:
            classroom = db.session.query(Classroom).filter_by(code=classroom_code).first()
            if classroom is None:
                messages.append("找不到對應教室")

        if not course_name:
            messages.append("缺少 course_name")

        weekday_raw = (row.get("weekday") or "").strip()
        try:
            weekday = int(weekday_raw)
            if weekday < 0 or weekday > 6:
                raise ValueError
        except ValueError:
            messages.append("weekday 需為 0-6")
            weekday = 0

        start_time_raw = (row.get("start_time") or "").strip()
        end_time_raw = (row.get("end_time") or "").strip()
        try:
            start_time = datetime.strptime(start_time_raw, "%H:%M").time()
        except ValueError:
            messages.append("start_time 格式需為 HH:MM")
            start_time = None

        try:
            end_time = datetime.strptime(end_time_raw, "%H:%M").time()
        except ValueError:
            messages.append("end_time 格式需為 HH:MM")
            end_time = None

        if start_time and end_time and start_time >= end_time:
            messages.append("結束時間需晚於開始時間")

        if is_active_raw not in {"", "0", "1", "true", "false", "yes", "no", "y", "n"}:
            messages.append("is_active 需為 true/false 或 1/0")

        if messages:
            errors.append(
                {
                    "row_number": str(index),
                    "classroom_code": classroom_code or "-",
                    "course_name": course_name or "-",
                    "messages": "、".join(messages),
                }
            )
            continue

        valid_rows.append(
            {
                "classroom_id": classroom.id,
                "classroom_code": classroom_code,
                "course_name": course_name,
                "instructor_name": instructor_name,
                "weekday": weekday,
                "start_time": start_time.strftime("%H:%M"),
                "end_time": end_time.strftime("%H:%M"),
                "semester_label": semester_label,
                "is_active": is_active_raw in {"1", "true", "yes", "y"},
            }
        )

    return valid_rows, errors


def _persist_schedule_rows(valid_rows: list[dict[str, str]]) -> int:
    imported = 0
    for row in valid_rows:
        schedule = CourseSchedule(
            classroom_id=row["classroom_id"],
            course_name=row["course_name"],
            instructor_name=row["instructor_name"] or None,
            weekday=row["weekday"],
            start_time=datetime.strptime(row["start_time"], "%H:%M").time(),
            end_time=datetime.strptime(row["end_time"], "%H:%M").time(),
            semester_label=row["semester_label"],
            is_active=bool(row["is_active"]),
        )
        db.session.add(schedule)
        imported += 1
    db.session.commit()
    return imported


def _build_csv_template_text() -> str:
    return (
        "classroom_code,course_name,instructor_name,weekday,start_time,end_time,semester_label,is_active\n"
        "L105,資料結構,王老師,0,09:00,11:00,114-2,true\n"
        "ED202,教學設計,陳老師,2,13:00,15:00,114-2,true\n"
    )


@bp.route("/")
@login_required
@roles_required("super_admin", "staff_manager", "workstudy_manager")
def dashboard():
    return render_template("admin/dashboard.html", **_get_dashboard_context())


@bp.route("/classrooms", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def upsert_classroom():
    code = request.form["code"].strip().upper()
    room = db.session.query(Classroom).filter_by(code=code).first()
    if room is None:
        room = Classroom(code=code)
        db.session.add(room)

    room.name = request.form["name"].strip()
    room.location = request.form["location"].strip()
    room.room_type = request.form.get("room_type", "").strip() or "教室"
    room.capacity = int(request.form["capacity"])
    room.is_online_bookable = request.form.get("is_online_bookable") == "on"
    booking_start_time = datetime.strptime(request.form["booking_start_time"], "%H:%M").time()
    booking_end_time = datetime.strptime(request.form["booking_end_time"], "%H:%M").time()
    if booking_start_time >= booking_end_time:
        flash("教室可借用結束時間需晚於開始時間。", "danger")
        return redirect(url_for("admin.dashboard"))
    room.booking_start_time = booking_start_time
    room.booking_end_time = booking_end_time
    room.note = request.form.get("note", "").strip() or None
    db.session.commit()
    flash("教室資料已更新。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/rules", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def upsert_rule():
    key = request.form["key"].strip()
    value = request.form["value"].strip()
    description = request.form["description"].strip()

    rule = db.session.query(SystemRule).filter_by(key=key).first()
    if rule is None:
        rule = SystemRule(key=key, value=value, description=description)
        db.session.add(rule)
    else:
        rule.value = value
        rule.description = description
    db.session.commit()
    flash("規則已更新。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/periods", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def replace_periods():
    raw_text = request.form.get("periods_text", "").strip()
    if not raw_text:
        flash("請提供至少一筆節次設定。", "danger")
        return redirect(url_for("admin.dashboard"))

    parsed_periods: list[BookingPeriod] = []
    seen_codes: set[str] = set()
    previous_end_time = None
    for line_number, line in enumerate(raw_text.splitlines(), start=1):
        stripped = line.strip()
        if not stripped:
            continue
        parts = [part.strip() for part in stripped.split(",")]
        if len(parts) != 3:
            flash(f"第 {line_number} 行格式錯誤，請使用 01,08:10,09:00。", "danger")
            return redirect(url_for("admin.dashboard"))

        code, start_raw, end_raw = parts
        if code in seen_codes:
            flash(f"第 {line_number} 行節次代碼重複：{code}。", "danger")
            return redirect(url_for("admin.dashboard"))
        seen_codes.add(code)

        try:
            start_time = datetime.strptime(start_raw, "%H:%M").time()
            end_time = datetime.strptime(end_raw, "%H:%M").time()
        except ValueError:
            flash(f"第 {line_number} 行時間格式錯誤，請使用 HH:MM。", "danger")
            return redirect(url_for("admin.dashboard"))

        if start_time >= end_time:
            flash(f"第 {line_number} 行結束時間需晚於開始時間。", "danger")
            return redirect(url_for("admin.dashboard"))
        if previous_end_time and start_time <= previous_end_time:
            flash(f"第 {line_number} 行開始時間需晚於上一節次結束時間。", "danger")
            return redirect(url_for("admin.dashboard"))
        previous_end_time = end_time

        parsed_periods.append(
            BookingPeriod(
                code=code,
                start_time=start_time,
                end_time=end_time,
                sort_order=line_number,
                is_active=True,
            )
        )

    db.session.query(BookingPeriod).delete()
    for period in parsed_periods:
        db.session.add(period)
    db.session.commit()
    flash("節次設定已更新。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/users/<int:user_id>/roles", methods=["POST"])
@login_required
@roles_required("super_admin")
def update_user_roles(user_id: int):
    user = db.session.get(User, user_id)
    if user is None:
        flash("找不到使用者。", "danger")
        return redirect(url_for("admin.dashboard"))

    selected_roles = request.form.getlist("roles")
    user.roles = db.session.query(Role).filter(Role.name.in_(selected_roles)).all()
    db.session.commit()
    flash("使用者角色已更新。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/manual-booking", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager", "workstudy_manager")
def manual_booking():
    requester = db.session.get(User, int(request.form["requester_id"]))
    classroom = db.session.get(Classroom, int(request.form["classroom_id"]))
    if requester is None or classroom is None:
        flash("借用對象或教室不存在。", "danger")
        return redirect(url_for("admin.dashboard"))

    start_at = datetime.strptime(request.form["start_at"], "%Y-%m-%dT%H:%M")
    end_at = datetime.strptime(request.form["end_at"], "%Y-%m-%dT%H:%M")

    try:
        booking = create_booking(
            requester=requester,
            classroom=classroom,
            title=request.form["title"].strip(),
            purpose=request.form["purpose"].strip(),
            attendee_count=int(request.form["attendee_count"]),
            start_at=start_at,
            end_at=end_at,
            source="manual",
        )
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin.dashboard"))

    if classroom.is_online_bookable:
        flash(f"人工登記完成（單號 #{booking.id}）。", "success")
    else:
        flash(f"已為不可線上借用教室完成人工登記（單號 #{booking.id}）。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/schedules", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def upsert_schedule():
    classroom = db.session.get(Classroom, int(request.form["classroom_id"]))
    if classroom is None:
        flash("找不到教室。", "danger")
        return redirect(url_for("admin.dashboard"))

    schedule_id = request.form.get("schedule_id", type=int)
    schedule = db.session.get(CourseSchedule, schedule_id) if schedule_id else None
    if schedule is None:
        schedule = CourseSchedule(classroom_id=classroom.id)
        db.session.add(schedule)

    start_time = datetime.strptime(request.form["start_time"], "%H:%M").time()
    end_time = datetime.strptime(request.form["end_time"], "%H:%M").time()
    if start_time >= end_time:
        flash("課表結束時間需晚於開始時間。", "danger")
        return redirect(url_for("admin.dashboard"))

    schedule.classroom_id = classroom.id
    schedule.course_name = request.form["course_name"].strip()
    schedule.instructor_name = request.form.get("instructor_name", "").strip() or None
    schedule.weekday = int(request.form["weekday"])
    schedule.start_time = start_time
    schedule.end_time = end_time
    schedule.semester_label = request.form.get("semester_label", "").strip() or "current"
    schedule.is_active = request.form.get("is_active") == "on"
    db.session.commit()
    flash("課表已更新。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/schedules/<int:schedule_id>/delete", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def delete_schedule(schedule_id: int):
    schedule = db.session.get(CourseSchedule, schedule_id)
    if schedule is None:
        flash("找不到課表。", "danger")
        return redirect(url_for("admin.dashboard"))

    db.session.delete(schedule)
    db.session.commit()
    flash("課表已刪除。", "info")
    return redirect(url_for("admin.dashboard"))


@bp.route("/schedules/import", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def import_schedules():
    preview_payload = request.form.get("preview_payload", "").strip()
    if preview_payload:
        try:
            valid_rows = json.loads(preview_payload)
        except json.JSONDecodeError:
            flash("匯入預覽資料已失效，請重新上傳檔案。", "danger")
            return redirect(url_for("admin.dashboard"))
        if not valid_rows:
            flash("沒有可匯入的資料列。", "danger")
            return redirect(url_for("admin.dashboard"))
        imported = _persist_schedule_rows(valid_rows)
        flash(f"課表匯入完成，新增 {imported} 筆。", "success")
        return redirect(url_for("admin.dashboard"))

    file = request.files.get("schedule_file")
    if file is None or not file.filename:
        flash("請選擇要匯入的課表檔案。", "danger")
        return redirect(url_for("admin.dashboard"))

    try:
        rows, fieldnames = _read_schedule_rows(file)
    except UnicodeDecodeError:
        flash("CSV 檔案編碼需為 UTF-8。", "danger")
        return redirect(url_for("admin.dashboard"))
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin.dashboard"))

    if not fieldnames or not REQUIRED_SCHEDULE_COLUMNS.issubset(set(fieldnames)):
        flash("匯入欄位不足，需包含 classroom_code, course_name, instructor_name, weekday, start_time, end_time, semester_label, is_active。", "danger")
        return redirect(url_for("admin.dashboard"))

    valid_rows, errors = _validate_schedule_rows(rows)
    if not valid_rows and not errors:
        flash("匯入檔案沒有可讀取的資料列。", "danger")
        return redirect(url_for("admin.dashboard"))

    context = _get_dashboard_context()
    context["schedule_import_preview"] = valid_rows
    context["schedule_import_errors"] = errors
    context["schedule_import_filename"] = file.filename
    context["schedule_import_payload"] = json.dumps(valid_rows, ensure_ascii=False)
    if errors:
        flash(f"預覽完成，可匯入 {len(valid_rows)} 筆，錯誤 {len(errors)} 筆。", "warning")
    else:
        flash(f"預覽完成，可匯入 {len(valid_rows)} 筆。", "info")
    return render_template("admin/dashboard.html", **context)


@bp.route("/schedules/template.csv")
@login_required
@roles_required("super_admin", "staff_manager")
def download_schedule_template():
    payload = _build_csv_template_text()
    return Response(
        payload,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=course-schedule-template.csv"},
    )


@bp.route("/schedules/template.xlsx")
@login_required
@roles_required("super_admin", "staff_manager")
def download_schedule_template_xlsx():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedules"
    for row in csv.reader(StringIO(_build_csv_template_text())):
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=course-schedule-template.xlsx"},
    )


@bp.route("/schedules/template.ods")
@login_required
@roles_required("super_admin", "staff_manager")
def download_schedule_template_ods():
    spreadsheet = OpenDocumentSpreadsheet()
    table = Table(name="Schedules")
    spreadsheet.spreadsheet.addElement(table)
    for values in csv.reader(StringIO(_build_csv_template_text())):
        row = TableRow()
        for value in values:
            cell = TableCell()
            cell.addElement(P(text=value))
            row.addElement(cell)
        table.addElement(row)
    buffer = BytesIO()
    spreadsheet.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.oasis.opendocument.spreadsheet",
        headers={"Content-Disposition": "attachment; filename=course-schedule-template.ods"},
    )


@bp.route("/blocks", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def upsert_block():
    classroom = db.session.get(Classroom, int(request.form["classroom_id"]))
    if classroom is None:
        flash("找不到教室。", "danger")
        return redirect(url_for("admin.dashboard"))

    start_at = datetime.strptime(request.form["start_at"], "%Y-%m-%dT%H:%M")
    end_at = datetime.strptime(request.form["end_at"], "%Y-%m-%dT%H:%M")
    if start_at >= end_at:
        flash("停用結束時間需晚於開始時間。", "danger")
        return redirect(url_for("admin.dashboard"))

    block = ClassroomBlock(
        classroom_id=classroom.id,
        title=request.form["title"].strip(),
        reason=request.form.get("reason", "").strip() or None,
        block_type=request.form.get("block_type", "maintenance").strip() or "maintenance",
        start_at=start_at,
        end_at=end_at,
        is_active=request.form.get("is_active") == "on",
    )
    db.session.add(block)
    impacted_bookings = (
        db.session.query(BookingRequest)
        .filter(
            BookingRequest.classroom_id == classroom.id,
            BookingRequest.status.in_(["approved", "pending_approval"]),
            BookingRequest.start_at < end_at,
            BookingRequest.end_at > start_at,
        )
        .all()
    )
    for booking in impacted_bookings:
        create_notification(
            user_id=booking.requester_id,
            title="借用時段受到停用影響",
            body=f"{classroom.code} {booking.title} 與停用時段《{block.title}》重疊，請儘快確認。",
            link=url_for("booking.list_bookings"),
        )
    db.session.commit()
    flash("教室停用時段已建立。", "success")
    return redirect(url_for("admin.dashboard"))


@bp.route("/blocks/<int:block_id>/delete", methods=["POST"])
@login_required
@roles_required("super_admin", "staff_manager")
def delete_block(block_id: int):
    block = db.session.get(ClassroomBlock, block_id)
    if block is None:
        flash("找不到停用時段。", "danger")
        return redirect(url_for("admin.dashboard"))

    db.session.delete(block)
    db.session.commit()
    flash("停用時段已刪除。", "info")
    return redirect(url_for("admin.dashboard"))
